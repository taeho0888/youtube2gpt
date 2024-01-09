import openpyxl
from youtube_transcript_api import YouTubeTranscriptApi

class YouTubeScriptExtractor:
    def get_youtube_id(self, youtube_link):
        try:
            return youtube_link.split("?v=")[1]
        except Exception as e:
            return None

    def extract_script(self, youtube_id):
        try:
            scripts = YouTubeTranscriptApi.get_transcript(youtube_id, languages=['ko'])
            return " ".join(script["text"] for script in scripts)
        except Exception as e:
            return "[오류] : 스크립트 추출 실패"

    def process_scripts(self, sheet, input_column, output_column):
        for row in sheet.iter_rows(min_row=3, min_col=input_column, max_col=input_column):
            if sheet.cell(row=row[0].row, column=output_column).value:
                continue

            youtube_link = row[0].value
            youtube_id = self.get_youtube_id(youtube_link)

            if youtube_id:
                content = self.extract_script(youtube_id)
            else:
                content = f"[오류] : {youtube_link}는 올바른 URL이 아닙니다."
                print(content)

            sheet.cell(row=row[0].row, column=output_column, value=content)
